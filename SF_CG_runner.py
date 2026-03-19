import json
import os
import sys
import shutil
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook

os.chdir(os.path.dirname(os.path.dirname(os.getcwd())))

from utils.format import *
from mst.model.SF_CG.SF_CG_config import *
from utils.compare_file import compare_excel
from utils.log_file import Timestamplog

file_keyword = LOG_DIR / "SF CG Runner"
sys.stdout = Timestamplog(file_keyword)

date = datetime.now().strftime("%Y%d")
time = datetime.now().strftime("%H%M%S")

for t1_timepoint in TIMEPOINTS:
    if OUTPUT_DIR_FINAL:
        current_run_folder = CG_OUTPUT_PATH
    else:
        current_run_folder = CG_OUTPUT_PATH / f"{date}_{time}"
    os.makedirs(current_run_folder, exist_ok=True)

    date = datetime.now().strftime("%Y%d")
    time = datetime.now().strftime("%H%M%S")

    for sce in SCENARIOS.keys():
        input_file = SCENARIOS[sce]
        p0_date = SCENARIOS_date[sce]
        config = {
            "scenario": f"{sce}",
            "input_file": f"{str(input_file)}",
            "p0_date": f"{p0_date}",
            "t1_timepoint": f"{t1_timepoint}",
            "crf": str(current_run_folder),
        }
        with open('./cg_config.json', 'w') as f:
            json.dump(config, f)
        with open('./cg_config.json', 'r') as f:
            config = json.load(f)
        exec(open('mst/model/SF_CG/SF_CG_impact.py').read())

    # Factor chart update
    factor_book = load_workbook(f'{current_run_folder}\\Chart\\Chart_SF_factor_impact.xlsx')
    factor_chart = factor_book['PasteResultHere']

    def get_header_index_map(ws):
        return {cell.value: idx for idx, cell in enumerate(ws[1])}

    factor_chart_col_index = get_header_index_map(factor_chart)
    scenario_first_row_idx = 2

    for tp in SCENARIOS.keys():
        tp_scenario = tp
        tp_factor_output = f"{tp_scenario}_{t1_timepoint}_factor_impact.csv"
        tp_factor_result = pd.read_csv(f'{current_run_folder}\\{tp_factor_output}')
        columns_to_copy = tp_factor_result.columns

        for i, row in tp_factor_result.iterrows():
            excel_row = 1 + scenario_first_row_idx + i
            for col in columns_to_copy:
                col_idx = factor_chart_col_index[col]
                factor_chart.cell(row=excel_row, column=col_idx).value = row[col]
            factor_chart.cell(row=excel_row, column=factor_chart_col_index['Scenario']).value = tp_scenario
            factor_chart.cell(row=excel_row, column=factor_chart_col_index['timepoint']).value = t1_timepoint
            factor = factor_chart.cell(row=excel_row, column=factor_chart_col_index['Factor']).value.replace("_", " ")
            factor_chart.cell(row=excel_row, column=factor_chart_col_index['Key']).value = (
                str(tp_scenario) + str(t1_timepoint) + str(factor))
            factor_chart.cell(row=excel_row, column=factor_chart_col_index['#']).value = 1

        factor_chart.delete_rows(excel_row + 1, factor_chart.max_row - excel_row)
        scenario_first_row_idx += tp_factor_result.shape[0]

    factor_book.save(f'{current_run_folder}\\Chart\\Chart_SF_factor_impact.xlsx')

    # CG chart update
    cg_chart_book = load_workbook(f'{current_run_folder}\\Chart\\SF_CG_impact_Chart.xlsx')
    cg_chart = cg_chart_book['PasteResultHere']
    cg_chart_col_index = get_header_index_map(cg_chart)
    scenario_first_row_idx = 2

    for tp in SCENARIOS.keys():
        tp_scenario = tp
        tp_cg_output = f"{tp_scenario}_{t1_timepoint}_keyTable.csv"
        tp_cg_result = pd.read_csv(f'{current_run_folder}\\{tp_cg_output}')
        columns_to_copy = tp_cg_result.columns[1:]

        for i, row in tp_cg_result.iterrows():
            excel_row = 1 + scenario_first_row_idx + i
            for col in columns_to_copy:
                col_idx = cg_chart_col_index[col]
                cg_chart.cell(row=excel_row, column=col_idx).value = row[col]

        cg_chart.delete_rows(excel_row + 1, cg_chart.max_row - excel_row)
        scenario_first_row_idx += tp_cg_result.shape[0]

    cg_chart_book.save(f'{current_run_folder}\\Chart\\SF_CG_impact_Chart.xlsx')
